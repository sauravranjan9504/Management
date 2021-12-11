import tkinter as tk
#from tkinter import *
from tkinter import ttk
from tkcalendar import *
import time
import pyttsx3
from tkinter import messagebox
import csv 
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import mysql.connector

root=tk.Tk()
root.title("student management system")
root.state("zoomed")
root["bg"]="skyblue"
#=================================================================================================================================
#----------------------------------------------common functions and variables started-----------------------------------------------------------
#=================================================================================================================================
wb=load_workbook("student.xlsx")
sheet=wb.active

#lab = tk.Label(root , text="Only for trial purpose",font="timesnewroman 10 bold underline" , bg="skyblue")
#lab.place(x=1200 , y=670)
#
#lab1= tk.Label(root , text="Rahul Patil",font="timesnewroman 10 underline" , bg="skyblue")
#lab1.place(x=1225 , y=700)

def excel():
    sheet.column_dimensions['A'].width=30
    sheet.column_dimensions['B'].width=10
    sheet.column_dimensions['C'].width=15
    sheet.column_dimensions['D'].width=15
    sheet.column_dimensions['E'].width=15
    sheet.column_dimensions['F'].width=15
    sheet.column_dimensions['G'].width=10
    sheet.column_dimensions['H'].width=15
    sheet.column_dimensions['I'].width=20
    sheet.column_dimensions['J'].width=20
    sheet.column_dimensions['K'].width=20
    sheet.column_dimensions['L'].width=20

    sheet.cell(row=1 , column=1).value = "Student name"
    sheet.cell(row=1 , column=2).value = "Admission year"
    sheet.cell(row=1 , column=3).value = "Date of birth"
    sheet.cell(row=1 , column=4).value = "Fees"
    sheet.cell(row=1 , column=5).value = "Standard"
    sheet.cell(row=1 , column=6).value = "Division"
    sheet.cell(row=1 , column=7).value = "Gender"
    sheet.cell(row=1 , column=8).value = "Cast"
    sheet.cell(row=1 , column=9).value = "Category"
    sheet.cell(row=1 , column=10).value = "Religion"
    sheet.cell(row=1 , column=11).value = "Mobile no."
    sheet.cell(row=1 , column=12).value = "Adhar no."    

excel()

def clock():
    hour = time.strftime("%I")
    minute = time.strftime("%M")
    second = time.strftime("%S")
    am_or_pm = time.strftime("%p")

    live_time = tk.Label(root , text = "" , fg="white", font = "timesnewroman 10 bold underline" , bg="black" , relief="sunken" , bd=3)
    live_time.place(x=1275,y=10)

    live_time.config(text = hour + ":" + minute + ":" + second  + " " +am_or_pm )
    live_time.after(1000 , clock)

speak = pyttsx3.init()
def for_speak(str):
    speak.setProperty("rate" , 150)
    speak.setProperty("volume" , 1)
    voices = speak.getProperty("voices")
    speak.setProperty("voice" , voices[1].id)
    speak.say(str)
    speak.runAndWait()

for_speak("hello")



#=================================================================================================================================
#----------------------------------------------common functions ended-------------------------------------------------------------
#=================================================================================================================================


#=================================================================================================================================
#----------------------------------------------Admission Section started---------------------------------------------------
#=================================================================================================================================

#function section
def student_admission():
    student_admission_labelframe = tk.LabelFrame(root , 
                                                text = "Admission Section ->" ,
                                                font = "timesnewroman 15 bold italic underline" , 
                                                relief = "ridge" ,
                                                bd = 5 , 
                                                bg = "#660033" , 
                                                fg = "white" ,
                                                height = 733 ,
                                                width = 1355)
    student_admission_labelframe.place(x=5 , y=5)

    clock()

    #back button 1st
    def back1():
        student_admission_labelframe.destroy()
        side_frame.destroy()

    back1 = tk.Button(student_admission_labelframe , 
                    text = "<--Back" , 
                    font = "timesnewroman 10 underline" ,
                    relief = "groove" , 
                    bd = 3 ,
                    bg = "black" , 
                    fg = "white" ,
                    height = 1 ,
                    activebackground = "white" ,
                    activeforeground = "black" ,
                    cursor = 'hand2' ,
                    command = back1)

    back1.place(x=5 , y=5)

    #admission frame 1

    entry_frame1 = tk.Frame(root , 
                            relief = "ridge" ,
                            bd = 5 , 
                            bg = "#ff0059" , 
                            height = 690 ,
                            width = 425)
    entry_frame1.place(x=15 , y=35)

    def go_bk():
        entry_frame1.destroy()
        student_admission_labelframe.destroy()

    bk_butt = tk.Button(entry_frame1 , text="<--Back" , command=go_bk , bg="black" , fg="white" , activebackground="yellow" , activeforeground="black",relief="sunken" , bd=3)
    bk_butt.place(x=5 , y=5)

    student_name_label = tk.Label(entry_frame1 , text = "Student name:-",font="timesnewroman 14" , relief = "flat" , bd=1 ,bg="black" , fg="white" ,height=1)   
    student_name_label.place(x=15 , y=55)
    
    student_admissionyear_label = tk.Label(entry_frame1 , text = "Admission date:-",font="timesnewroman 14" , relief = "flat" , bd=1 ,bg="black" , fg="white" ,height=1)    
    student_admissionyear_label.place(x=15 , y=100)
  
    student_dateofbirth_label = tk.Label(entry_frame1 , text = "Date of birth:-",font="timesnewroman 14" , relief = "flat" , bd=1 ,bg="black" , fg="white" ,height=1)    
    student_dateofbirth_label.place(x=15 , y=145)

    student_fees_label = tk.Label(entry_frame1 , text = "fees:-",font="timesnewroman 14" , relief = "flat" , bd=1 ,bg="black" , fg="white" ,height=1)    
    student_fees_label.place(x=15 , y=190)

    student_standard_label = tk.Label(entry_frame1 , text = "standard:-",font="timesnewroman 14" , relief = "flat" , bd=1 ,bg="black" , fg="white" ,height=1)    
    student_standard_label.place(x=15 , y=235)
    
    student_Division_label = tk.Label(entry_frame1 , text = "Division:-",font="timesnewroman 14" , relief = "flat" , bd=1 ,bg="black" , fg="white" ,height=1)    
    student_Division_label.place(x=15 , y=280)

    student_gender_label = tk.Label(entry_frame1 , text = "Gender:-",font="timesnewroman 14" , relief = "flat" , bd=1 ,bg="black" , fg="white" ,height=1)    
    student_gender_label.place(x=15, y=325)

    student_cast_label = tk.Label(entry_frame1 , text = "Cast:-",font="timesnewroman 14" , relief = "flat" , bd=1 ,bg="black" , fg="white" ,height=1)    
    student_cast_label.place(x=15 , y=370)
    
    student_category_label = tk.Label(entry_frame1 , text = "Category:-",font="timesnewroman 14" , relief = "flat" , bd=1 ,bg="black" , fg="white" ,height=1)    
    student_category_label.place(x=15 , y=415)

    student_religion_label =tk.Label(entry_frame1 , text="religion:-" , font="timesnewroman 14" ,relief = "flat" , bd=1 ,bg="black" , fg="white" ,height=1)
    student_religion_label.place(x=15 , y=460)

    student_mobno_label =tk.Label(entry_frame1 , text="mobno:-" , font="timesnewroman 14" ,relief = "flat" , bd=1 ,bg="black" , fg="white" ,height=1)
    student_mobno_label.place(x=15 , y=505)
  
    student_adharno_label =tk.Label(entry_frame1 , text="adharno:-" , font="timesnewroman 14" ,relief = "flat" , bd=1 ,bg="black" , fg="white" ,height=1)
    student_adharno_label.place(x=15 , y=555)


# religion, adhar no , omb no. , std 
    
    #entries

    #name
    student_name_entry = tk.Entry(entry_frame1 , relief = "sunken" , bd=3 )    
    student_name_entry.place(x=250 , y=55)
    student_name_entry.focus_set()

    #admission year
    student_admissionyear_entry = DateEntry(entry_frame1 , font='timesnewroman 10 bold',relief = "sunken" , bd=3 )    
    student_admissionyear_entry.place(x=250 , y=100)
    student_admissionyear_entry.focus_set()

    #date of birth
    student_dateofbirth_entry = DateEntry(entry_frame1 ,font="timesnewroman 10 bold ", relief = "sunken" , bd=3 )    
    student_dateofbirth_entry.place(x=250 , y=145)
    student_dateofbirth_entry.focus_set()

    #fees
    student_fees_entry = tk.Entry(entry_frame1 , relief = "sunken" , bd=3 )    
    student_fees_entry.place(x=250 , y=190)
    student_fees_entry.focus_set()

    #standard
    studentstandard = tk.StringVar()
    studentstandard.set("--select--")
    student_standard_entry = tk.OptionMenu(entry_frame1  , studentstandard,"1 st" , "2 nd" , "3 rd", "4 th", "5 th", "6 th" ,"7 th" ,"8 th" ,"9 th" ,"10 th", "11 th", "12 th")    
    student_standard_entry.config(font="timesnewroman 10 " , relief="raised" ,bd=3 ,bg="white")
    student_standard_entry.place(x=250 , y=235)
    student_standard_entry.focus_set()

    #division
    studentdivision = tk.StringVar()
    student_Division_entry = tk.Spinbox(entry_frame1 ,values = ("A" , "B" , "C" ,"D" , "E" ,"F" ,"G") ,relief = "sunken" , bd=3 )    
    studentdivision.set("--select--")
    student_Division_entry.place(x=250 , y=280)
    student_Division_entry.focus_set()

    #gender
    studentgender = tk.IntVar()
    student_gender_entry = tk.Radiobutton(entry_frame1 ,text = "Male" , relief = "sunken" , bd=3 ,variable=studentgender, value = 1)    
    student_gender_entry.place(x=250 , y=325) 
    student_gender_entry = tk.Radiobutton(entry_frame1 ,text="Female", relief = "sunken" , bd=3 ,variable=studentgender, value=2)    
    student_gender_entry.place(x=320 , y=325)
    student_gender_entry.focus_set()

    #cast
    student_cast_entry = tk.Entry(entry_frame1 , relief = "sunken" , bd=3)    
    student_cast_entry.place(x=250 , y=370)
    student_cast_entry.focus_set()

    #category
    student_category_entry = tk.Entry(entry_frame1 , relief = "sunken" , bd=3 )    
    student_category_entry.place(x=250 , y=415)
    student_category_entry.focus_set()

    #religion
    student_religion_entry = tk.Entry(entry_frame1 , relief = "sunken" , bd=3 )
    student_religion_entry.place(x=250 , y=460)
   
    #mobile number
    student_mobno_entry = tk.Entry(entry_frame1 , relief = "sunken" , bd=3 )
    student_mobno_entry.place(x=250 , y=505)

    #adhar no.
    student_adharno_entry = tk.Entry(entry_frame1,relief = "sunken" , bd=3 )
    student_adharno_entry.place(x=250 , y=555)

    def admission():
        if (student_name_entry.get() == ""):
            messagebox.showerror("error" , "All fields are required.")
        elif (student_admissionyear_entry.get() == "" ):
            messagebox.showerror("error" , "All fields are required.")
        elif (student_dateofbirth_entry.get() == "" ):
            messagebox.showerror("error" , "All fields are required.")
        elif (student_fees_entry.get() == ""):
            messagebox.showerror("error" , "All fields are required.")
        elif (studentstandard.get() == "--select--"): 
            messagebox.showerror("error" , "All fields are required.")
        elif (studentdivision.get() == "" ):
            messagebox.showerror("error" , "All fields are required.")
        elif (student_cast_entry.get() == "" ): 
            messagebox.showerror("error" , "All fields are required.")
        elif (student_category_entry.get() == ""):
            messagebox.showerror("error" , "All fields are required.")
        else:

            #==========Data base========
            mydb=mysql.connector.connect(host="localhost" , user="root" , password="" , database="student_information")
            mycursor=mydb.cursor()
            #mycursor.execute("CREATE TABLE buses (bus_number VARCHAR(255) , from VARCHAR(255) , to VARCHAR(255) , arrival_time VARCHAR(255) , departure_time VARCHAR(255))")
            sql="INSERT INTO studentdata (student_name , admission_date , dob , fees , std , division , gender , student_cast , category , religion , mobile_number , adhar_number) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
            val=(student_name_entry.get() , student_admissionyear_entry.get() , student_dateofbirth_entry.get() ,student_fees_entry.get() ,studentstandard.get() , student_Division_entry.get() ,studentgender.get() ,student_cast_entry.get() , student_category_entry.get() ,student_religion_entry.get() ,student_mobno_entry.get(),student_adharno_entry.get())
            mycursor.execute(sql,val)
            mydb.commit()
            mydb.close()





            current_row=sheet.max_row
            #current_column = sheet.max_column
            
            sheet.cell(row=current_row + 1, column=1).value = student_name_entry.get() 
            sheet.cell(row=current_row + 1, column=2).value = student_admissionyear_entry.get() 
            sheet.cell(row=current_row + 1, column=3).value = student_dateofbirth_entry.get()
            sheet.cell(row=current_row + 1, column=4).value = student_fees_entry.get() 
            sheet.cell(row=current_row + 1, column=5).value = studentstandard.get() 
            sheet.cell(row=current_row + 1, column=6).value = student_Division_entry.get() 
            if studentgender.get()==1:
                sheet.cell(row=current_row + 1, column=7).value = "M" 
            else:
                sheet.cell(row=current_row + 1, column=7).value = "F"
            sheet.cell(row=current_row + 1, column=8).value = student_cast_entry.get() 
            sheet.cell(row=current_row + 1, column=9).value = student_category_entry.get() 
            sheet.cell(row=current_row + 1, column=10).value= student_religion_entry.get()
            sheet.cell(row=current_row + 1, column=11).value= student_mobno_entry.get()
            sheet.cell(row=current_row + 1, column=12).value= student_adharno_entry.get()

            wb.save("student.xlsx")


            messagebox.showinfo("congrates" , "your response is successfully recorded. \n Thank you , Use it again.")

            student_name_entry.delete(0,len(student_name_entry.get()))
            student_admissionyear_entry.delete(0,len(student_admissionyear_entry.get()))
            student_dateofbirth_entry.delete(0,len(student_dateofbirth_entry.get()))
            student_fees_entry.delete(0,len(student_fees_entry.get()))
            
            student_Division_entry.delete(0,len(student_Division_entry.get()))
            student_cast_entry.delete(0,len(student_cast_entry.get()))
            student_category_entry.delete(0,len(student_category_entry.get()))


 


    frame1_submit_button = tk.Button(entry_frame1 , text = "submit" , font="timesnewroman 10 bold" , relief = "sunken" , bd=3 , bg="black" , fg="white" , activebackground="white" , activeforeground="black" , cursor="hand2" , command = admission)    
    frame1_submit_button.place(x=350, y=600)

#====================================================================================
#------------------------------first frame completed---------------------------------
#====================================================================================

    side_frame = tk.LabelFrame(student_admission_labelframe , text="Previous records",
                                                font = "timesnewroman 10 bold italic underline" , 
                                                relief = "ridge" ,
                                                bd = 5 , 
                                                bg = "#ff0059" , 
                                                fg = "white" ,
                                                height = 690 ,
                                                width = 1000)
    side_frame.place(x=435 , y=3)

    #-------treeview--------

    tree = ttk.Treeview(side_frame , show="headings" , columns = ("Name","Admission year","DOB","Fees","STD","DIV","Gender","Cast","Category","Religion","Mob no.","Adhar no."), height=32)
    tree.pack(fill="both")   

    tree.column("Name",width=100)
    tree.column("Admission year",width=90)  
    tree.column("DOB",width=70) 
    tree.column("Fees",width=60) 
    tree.column("STD",width=50) 
    tree.column("DIV",width=50) 
    tree.column("Gender",width=50) 
    tree.column("Cast",width=65) 
    tree.column("Category",width=90)
    tree.column("Religion",width=85)
    tree.column("Mob no.",width=90)
    tree.column("Adhar no.",width=120)

    tree.heading("Name",text="Name")
    tree.heading("Admission year",text="Admission year")  
    tree.heading("DOB",text="DOB") 
    tree.heading("Fees",text="Fees") 
    tree.heading("STD",text="STD") 
    tree.heading("DIV",text="DIV") 
    tree.heading("Gender",text="Gender") 
    tree.heading("Cast",text="Cast") 
    tree.heading("Category",text="Category") 
    tree.heading("Religion",text="Religion")
    tree.heading("Mob no.",text="Mob no.")   
    tree.heading("Adhar no.",text="Adhar no.")

    hs=ttk.Scrollbar(side_frame ,orient="vertical",command=tree.yview)
    #hs.pack(side="right",fill="y")
    hs.place(x=890,y=20,height=700)
    tree.configure(yscrollcommand=hs.set)
#===========================pending work=============================================================
    def refresh():
        for i in tree.get_children():
            tree.delete(i)
        fp = pd.read_excel("student.xlsx") # Read xlsx file
        for i in range(len(fp.index.values)): # use for loop to get values in each line, _ is the number of line.
            tree.insert('','end',value=tuple(fp.iloc[i,[0,1,2,3,4,5,6,7,8,9,10,11]].values))   

#=======================pending work ended=============================================================

    ref_but = tk.Button(side_frame , text = "REF" ,width=2, command=refresh , foreground='white' , background='black',activeforeground="black" , activebackground='white')
    ref_but.place(x=0 , y=0)














#widget section
student_admission_button = tk.Button(root , 
                                    text = "Student admission" , 
                                    font = "timesnewroman 10 " , 
                                    relief = "sunken" , 
                                    bd = 3 , 
                                    cursor = "hand2" , 
                                    bg = "black" , 
                                    fg = "white" , 
                                    activebackground = "white" , 
                                    activeforeground = "black" ,
                                    command = student_admission)
student_admission_button.place(x=20 , y=20)


#=================================================================================================================================
#----------------------------------------------Admission Section ended------------------------------------------------------------
#=================================================================================================================================
root.mainloop()